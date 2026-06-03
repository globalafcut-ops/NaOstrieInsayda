import asyncio
import logging
import sqlite3
import traceback
import os
from datetime import datetime, timedelta, timezone

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from telegram import (
    Update,
    InlineKeyboardButton,
    InlineKeyboardMarkup,
    ReplyKeyboardMarkup,
    KeyboardButton,
    ReplyKeyboardRemove,
)
from telegram.ext import (
    Application,
    CommandHandler,
    MessageHandler,
    CallbackQueryHandler,
    ContextTypes,
    filters,
)

# =========================
# CONFIG
# =========================
BOT_TOKEN = "8694819008:AAHez9zAlV2DWV-j268-ijg5h1pGG0uO9"
MAIN_ADMIN_ID = 8587056338

REVIEWS_URL = "https://t.me/otzuvufinanceradar"
CHANNEL_URL = "https://t.me/+_WZ_tWObnHhjNzhi"

DB_FILE = "finance_radar.sqlite3"

STEP1_IMAGE = "step1.jpg"
STEP2_IMAGE = "step2.jpg"
STEP3_IMAGE = "step3.jpg"
STEP4_IMAGE = "step4.jpg"
FINAL_IMAGE = "final.jpg"

REMINDER_FIRST_AFTER = timedelta(hours=1)
REMINDER_SECOND_AFTER = timedelta(hours=3)
REMINDER_DAILY_AFTER = timedelta(days=1)
REMINDER_DAILY_UNTIL = timedelta(days=7)
REMINDER_RECHECK_EVERY_SECONDS = 600

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
)
logger = logging.getLogger(__name__)

# =========================
# DB
# =========================
class Database:
    def __init__(self, db_file: str):
        self.db_file = db_file
        self._init_db()

    def _conn(self):
        return sqlite3.connect(self.db_file)

    def _init_db(self):
        with self._conn() as conn:
            cur = conn.cursor()
            cur.execute("""
            CREATE TABLE IF NOT EXISTS users (
                user_id INTEGER PRIMARY KEY,
                username TEXT,
                first_name TEXT,
                name TEXT,
                experience TEXT,
                phone TEXT,
                source TEXT,
                registered INTEGER DEFAULT 0,
                step INTEGER DEFAULT 0,
                created_at TEXT DEFAULT CURRENT_TIMESTAMP,
                updated_at TEXT DEFAULT CURRENT_TIMESTAMP,
                reminder_stage INTEGER DEFAULT 0,
                last_reminder_at TEXT
            )
            """)
            cur.execute("""
            CREATE TABLE IF NOT EXISTS admins (
                user_id INTEGER PRIMARY KEY,
                added_by INTEGER,
                created_at TEXT DEFAULT CURRENT_TIMESTAMP
            )
            """)
            cur.execute("""
            CREATE TABLE IF NOT EXISTS logs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER,
                actor_id INTEGER,
                action TEXT,
                details TEXT,
                created_at TEXT DEFAULT CURRENT_TIMESTAMP
            )
            """)
            cur.execute("""
            CREATE TABLE IF NOT EXISTS errors (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                error TEXT,
                traceback TEXT,
                created_at TEXT DEFAULT CURRENT_TIMESTAMP
            )
            """)
            conn.commit()
        self.add_admin(MAIN_ADMIN_ID, added_by=MAIN_ADMIN_ID)

    def log_action(self, user_id, actor_id, action, details=""):
        with self._conn() as conn:
            conn.execute(
                "INSERT INTO logs (user_id, actor_id, action, details) VALUES (?, ?, ?, ?)",
                (user_id, actor_id, action, details),
            )
            conn.commit()

    def log_error(self, error, tb):
        with self._conn() as conn:
            conn.execute(
                "INSERT INTO errors (error, traceback) VALUES (?, ?)",
                (error, tb),
            )
            conn.commit()

    def upsert_user(self, user_id, username="", first_name="", source=None):
        with self._conn() as conn:
            cur = conn.cursor()
            cur.execute("SELECT user_id FROM users WHERE user_id = ?", (user_id,))
            exists = cur.fetchone() is not None

            if exists:
                if source is not None:
                    conn.execute("""
                    UPDATE users
                    SET username=?, first_name=?, source=COALESCE(source, ?), updated_at=CURRENT_TIMESTAMP
                    WHERE user_id=?
                    """, (username, first_name, source, user_id))
                else:
                    conn.execute("""
                    UPDATE users
                    SET username=?, first_name=?, updated_at=CURRENT_TIMESTAMP
                    WHERE user_id=?
                    """, (username, first_name, user_id))
            else:
                conn.execute("""
                INSERT INTO users (user_id, username, first_name, source, updated_at)
                VALUES (?, ?, ?, ?, CURRENT_TIMESTAMP)
                """, (user_id, username, first_name, source))

            conn.commit()

    def get_user_row(self, user_id):
        with self._conn() as conn:
            cur = conn.cursor()
            cur.execute("SELECT * FROM users WHERE user_id = ?", (user_id,))
            return cur.fetchone()

    def get_step(self, user_id):
        row = self.get_user_row(user_id)
        if not row:
            return 0
        return row[8]

    def set_step(self, user_id, step):
        with self._conn() as conn:
            conn.execute(
                "UPDATE users SET step=?, updated_at=CURRENT_TIMESTAMP WHERE user_id=?",
                (step, user_id),
            )
            conn.commit()

    def set_name(self, user_id, name):
        with self._conn() as conn:
            conn.execute(
                "UPDATE users SET name=?, updated_at=CURRENT_TIMESTAMP WHERE user_id=?",
                (name, user_id),
            )
            conn.commit()

    def set_experience(self, user_id, experience):
        with self._conn() as conn:
            conn.execute(
                "UPDATE users SET experience=?, updated_at=CURRENT_TIMESTAMP WHERE user_id=?",
                (experience, user_id),
            )
            conn.commit()

    def set_phone(self, user_id, phone):
        with self._conn() as conn:
            conn.execute(
                "UPDATE users SET phone=?, registered=1, updated_at=CURRENT_TIMESTAMP WHERE user_id=?",
                (phone, user_id),
            )
            conn.commit()

    def is_registered(self, user_id):
        row = self.get_user_row(user_id)
        return bool(row and row[7] == 1)

    def is_admin(self, user_id):
        with self._conn() as conn:
            cur = conn.cursor()
            cur.execute("SELECT 1 FROM admins WHERE user_id = ?", (user_id,))
            return cur.fetchone() is not None

    def add_admin(self, user_id, added_by=None):
        with self._conn() as conn:
            conn.execute(
                "INSERT OR IGNORE INTO admins (user_id, added_by) VALUES (?, ?)",
                (user_id, added_by),
            )
            conn.commit()

    def remove_admin(self, user_id):
        if user_id == MAIN_ADMIN_ID:
            return False
        with self._conn() as conn:
            conn.execute("DELETE FROM admins WHERE user_id = ?", (user_id,))
            conn.commit()
        return True

    def list_admins(self):
        with self._conn() as conn:
            cur = conn.cursor()
            cur.execute("SELECT user_id, added_by, created_at FROM admins ORDER BY user_id")
            return cur.fetchall()

    def stats(self):
        with self._conn() as conn:
            cur = conn.cursor()
            cur.execute("SELECT COUNT(*) FROM users")
            total = cur.fetchone()[0]
            cur.execute("SELECT COUNT(*) FROM users WHERE registered=1")
            done = cur.fetchone()[0]
            cur.execute("SELECT COUNT(*) FROM users WHERE registered=0 AND step>0")
            pending = cur.fetchone()[0]
            return total, done, pending

    def get_reminder_candidates(self):
        with self._conn() as conn:
            cur = conn.cursor()
            cur.execute("""
                SELECT user_id, name, first_name, step, registered, created_at, reminder_stage, last_reminder_at
                FROM users
                WHERE registered = 0 AND step > 0 AND step < 5
            """)
            return cur.fetchall()

    def set_reminder_stage(self, user_id, stage, last_reminder_at=None):
        with self._conn() as conn:
            conn.execute(
                "UPDATE users SET reminder_stage=?, last_reminder_at=?, updated_at=CURRENT_TIMESTAMP WHERE user_id=?",
                (stage, last_reminder_at, user_id),
            )
            conn.commit()

    def get_all_admins(self):
        with self._conn() as conn:
            cur = conn.cursor()
            cur.execute("SELECT user_id FROM admins")
            return [row[0] for row in cur.fetchall()]

db = Database(DB_FILE)

# =========================
# TEXTS
# =========================
STEP1_TEXT = """Привет, {name}! Это НА ОСТРИЕ ИНСАЙДА.

Сейчас идёт набор в приватную группу.

Мы открываем набор только когда рынок даёт реальную возможность заработать. Сейчас — такой момент.

📈 Прошлая сессия длилась 7 дней — участники зафиксировали прибыль в диапазоне от +16% до +22%.

💬 300+ участников уже поделились своими результатами

Работа проходит в приватном Telegram-канале:

🔹 даём точки входа;
🔹 ведём позиции вместе с участниками;
🔹 объясняем логику каждого решения;
🔹 контролируем риски;
🔹 фиксируем результаты по ходу работы.

Не жди следующего окна — забирай доступ прямо сейчас"""

STEP2_TEXT = "Давайте для начала познакомимся, напишите ваше имя ✍️"

STEP3_TEXT = "{name}, вы уже пробовали зарабатывать на финансовом рынке?"

STEP4_TEXT = "{name}, места в группе ограничены. Оставьте номер — менеджер свяжется с вами для подтверждения вашего места"

FINAL_TEXT = """✅ Поздравляем, {name}! Место зарезервировано. Наш менеджер свяжется с вами в ближайшее время для подтверждения 📞"""

ALREADY_REGISTERED_TEXT = "✅ Вы уже оставляли заявку. Наш менеджер скоро свяжется с вами для подтверждения 📞"

REMINDER_1_TEXT = """👋 {name}, вы начали регистрацию, но пока не завершили её.

Чтобы закрепить место в группе, просто продолжите анкету ✍️
Это займёт всего пару минут ✨"""

REMINDER_2_TEXT = """⏰ {name}, напоминаем: ваша регистрация ещё не завершена.

Место в группе остаётся за вами, но нужно закончить анкету, чтобы менеджер смог подтвердить заявку ✅
Это займёт ещё несколько минут 📩"""

REMINDER_3_TEXT = """🔔 {name}, ваша заявка всё ещё ожидает завершения.

Если хотите сохранить место, просто вернитесь к анкете и завершите регистрацию 🌟"""

# =========================
# KEYBOARDS
# =========================
def step1_keyboard():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("🔥 Получить доступ 🔥", callback_data="get_access")],
        [InlineKeyboardButton("💬 Отзывы участников 💬", url=REVIEWS_URL)],
    ])

def step3_keyboard():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("✅ Да, я успешно торгую на рынке", callback_data="exp_1")],
        [InlineKeyboardButton("📊 Есть небольшой опыт", callback_data="exp_2")],
        [InlineKeyboardButton("👂 Слышал от знакомых, хочу попробовать", callback_data="exp_3")],
        [InlineKeyboardButton("❌ Нет, я полный новичок", callback_data="exp_4")],
    ])

def phone_keyboard():
    return ReplyKeyboardMarkup(
        [[KeyboardButton("📱 Поделиться номером", request_contact=True)]],
        resize_keyboard=True,
        one_time_keyboard=True,
    )

def final_keyboard():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("💬 Отзывы участников 💬", url=REVIEWS_URL)],
        [InlineKeyboardButton("📢 Перейти в канал 📢", url=CHANNEL_URL)],
    ])

def continue_keyboard():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("📩 Продолжить регистрацию 📩", callback_data="continue_registration")]
    ])

# =========================
# HELPERS
# =========================
def get_display_name(update: Update):
    u = update.effective_user
    return u.first_name or u.username or ""

async def log_exception(exc: Exception):
    tb = traceback.format_exc()
    logger.exception(exc)
    db.log_error(str(exc), tb)

async def safe_delete_message(bot, chat_id, message_id):
    try:
        await bot.delete_message(chat_id=chat_id, message_id=message_id)
    except Exception:
        pass

async def delete_step_messages(context: ContextTypes.DEFAULT_TYPE):
    """Удаляет предыдущие сообщения с эффектом ✅ (для inline-сообщений)"""
    ids = context.user_data.get("step_messages", [])
    chat_id = context.user_data.get("chat_id")
    if chat_id and ids:
        for mid in ids:
            try:
                await context.bot.edit_message_caption(
                    chat_id=chat_id, message_id=mid, caption="✅"
                )
            except Exception:
                try:
                    await context.bot.edit_message_text(
                        chat_id=chat_id, message_id=mid, text="✅"
                    )
                except Exception:
                    pass
            await asyncio.sleep(0.3)
            await safe_delete_message(context.bot, chat_id, mid)
    context.user_data["step_messages"] = []

async def delete_step_messages_silent(context: ContextTypes.DEFAULT_TYPE):
    """Удаляет предыдущие сообщения без эффекта (для шага с ReplyKeyboard)"""
    ids = context.user_data.get("step_messages", [])
    chat_id = context.user_data.get("chat_id")
    if chat_id and ids:
        for mid in ids:
            await safe_delete_message(context.bot, chat_id, mid)
    context.user_data["step_messages"] = []

async def notify_admins(context: ContextTypes.DEFAULT_TYPE, text: str):
    """Отправляет уведомление всем админам"""
    admins = db.get_all_admins()
    for admin_id in admins:
        try:
            await context.bot.send_message(chat_id=admin_id, text=text)
        except Exception:
            pass

def file_exists(filename: str) -> bool:
    """Проверяет, существует ли файл"""
    return os.path.isfile(filename)

async def send_step_message(context, chat_id, image_file, caption, reply_markup):
    """
    Отправляет сообщение с изображением, если оно существует.
    Если изображения нет - отправляет текстовое сообщение.
    """
    try:
        if file_exists(image_file):
            with open(image_file, "rb") as photo:
                msg = await context.bot.send_photo(
                    chat_id=chat_id,
                    photo=photo,
                    caption=caption,
                    reply_markup=reply_markup,
                )
        else:
            msg = await context.bot.send_message(
                chat_id=chat_id,
                text=caption,
                reply_markup=reply_markup,
            )
        return msg
    except Exception as e:
        logger.error(f"Error sending step message: {e}")
        msg = await context.bot.send_message(
            chat_id=chat_id,
            text=caption,
            reply_markup=reply_markup,
        )
        return msg

# =========================
# HANDLERS
# =========================
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        user = update.effective_user
        source = context.args[0] if context.args else None
        db.upsert_user(user.id, user.username or "", user.first_name or "", source=source)

        if db.is_registered(user.id):
            await update.message.reply_text(ALREADY_REGISTERED_TEXT)
            return

        db.set_step(user.id, 1)
        name = get_display_name(update)

        await delete_step_messages(context)

        msg = await send_step_message(
            context,
            update.effective_chat.id,
            STEP1_IMAGE,
            STEP1_TEXT.format(name=name),
            step1_keyboard()
        )
        context.user_data["step_messages"] = [msg.message_id]
        context.user_data["chat_id"] = update.effective_chat.id

    except Exception as e:
        await log_exception(e)

async def on_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        query = update.callback_query
        await query.answer()
        user = update.effective_user
        db.upsert_user(user.id, user.username or "", user.first_name or "")
        data = query.data

        if data == "get_access":
            db.set_step(user.id, 2)
            await delete_step_messages(context)

            msg = await send_step_message(
                context,
                query.message.chat_id,
                STEP2_IMAGE,
                STEP2_TEXT,
                None
            )
            context.user_data["step_messages"] = [msg.message_id]
            context.user_data["chat_id"] = query.message.chat_id

        elif data.startswith("exp_"):
            if db.get_step(user.id) < 3:
                return

            mapping = {
                "exp_1": "✅ Да, я успешно торгую на рынке",
                "exp_2": "📊 Есть небольшой опыт",
                "exp_3": "👂 Слышал от знакомых, хочу попробовать",
                "exp_4": "❌ Нет, я полный новичок",
            }
            experience = mapping[data]
            db.set_experience(user.id, experience)
            db.set_step(user.id, 4)

            row = db.get_user_row(user.id)
            name = row[3] or get_display_name(update)

            await delete_step_messages(context)

            msg = await send_step_message(
                context,
                query.message.chat_id,
                STEP4_IMAGE,
                STEP4_TEXT.format(name=name),
                phone_keyboard()
            )
            context.user_data["step_messages"] = [msg.message_id]
            context.user_data["chat_id"] = query.message.chat_id

        elif data == "continue_registration":
            row = db.get_user_row(user.id)
            step = db.get_step(user.id)
            name = (row[3] if row else None) or (row[2] if row else None) or get_display_name(update)

            await delete_step_messages(context)

            if step == 2:
                msg = await send_step_message(
                    context,
                    query.message.chat_id,
                    STEP2_IMAGE,
                    STEP2_TEXT,
                    None
                )
                context.user_data["step_messages"] = [msg.message_id]
                context.user_data["chat_id"] = query.message.chat_id

            elif step == 3:
                msg = await send_step_message(
                    context,
                    query.message.chat_id,
                    STEP3_IMAGE,
                    STEP3_TEXT.format(name=name),
                    step3_keyboard()
                )
                context.user_data["step_messages"] = [msg.message_id]
                context.user_data["chat_id"] = query.message.chat_id

            elif step >= 4:
                msg = await send_step_message(
                    context,
                    query.message.chat_id,
                    STEP4_IMAGE,
                    STEP4_TEXT.format(name=name),
                    phone_keyboard()
                )
                context.user_data["step_messages"] = [msg.message_id]
                context.user_data["chat_id"] = query.message.chat_id

    except Exception as e:
        await log_exception(e)

async def on_text(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        user = update.effective_user
        text = (update.message.text or "").strip()
        db.upsert_user(user.id, user.username or "", user.first_name or "")
        step = db.get_step(user.id)

        if step == 2:
            db.set_name(user.id, text)
            db.set_step(user.id, 3)

            await delete_step_messages(context)

            msg = await send_step_message(
                context,
                update.effective_chat.id,
                STEP3_IMAGE,
                STEP3_TEXT.format(name=text),
                step3_keyboard()
            )
            context.user_data["step_messages"] = [msg.message_id]
            context.user_data["chat_id"] = update.effective_chat.id

    except Exception as e:
        await log_exception(e)

async def on_contact(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        user = update.effective_user
        contact = update.message.contact
        step = db.get_step(user.id)
        if step < 4:
            return

        db.set_phone(user.id, contact.phone_number)
        db.set_step(user.id, 5)
        db.set_reminder_stage(user.id, 99, None)
        row = db.get_user_row(user.id)
        name = row[3] or get_display_name(update)

        # Тихое удаление шага с телефоном
        await delete_step_messages_silent(context)

        # Убираем кнопку "Поделиться номером"
        msg_remove = await update.message.reply_text(".", reply_markup=ReplyKeyboardRemove())
        await safe_delete_message(context.bot, update.effective_chat.id, msg_remove.message_id)

        msg = await send_step_message(
            context,
            update.effective_chat.id,
            FINAL_IMAGE,
            FINAL_TEXT.format(name=name),
            final_keyboard()
        )
        context.user_data["step_messages"] = [msg.message_id]
        context.user_data["chat_id"] = update.effective_chat.id

        db.log_action(user.id, user.id, "registration_completed", f"phone={contact.phone_number}")

        # Уведомление всем админам
        username = f"@{user.username}" if user.username else "—"
        experience = row[4] or "—"
        source = row[6] or "—"
        notify_text = (
            f"🎉 Новая заявка!\n\n"
            f"👤 Имя: {name}\n"
            f"📞 Телефон: {contact.phone_number}\n"
            f"🔗 Telegram: {username}\n"
            f"📊 Опыт: {experience}\n"
            f"📍 Источник: {source}"
        )
        await notify_admins(context, notify_text)

    except Exception as e:
        await log_exception(e)

async def reminder(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        user = update.effective_user
        db.upsert_user(user.id, user.username or "", user.first_name or "")
        row = db.get_user_row(user.id)
        name = (row[3] if row else None) or get_display_name(update)
        await update.message.reply_text(
            REMINDER_1_TEXT.format(name=name),
            reply_markup=continue_keyboard(),
        )
    except Exception as e:
        await log_exception(e)

# =========================
# REMINDERS
# =========================
async def reminder_job(context: ContextTypes.DEFAULT_TYPE):
    now = datetime.now(timezone.utc)
    try:
        candidates = db.get_reminder_candidates()
        for user_id, name, first_name, step, registered, created_at, reminder_stage, last_reminder_at in candidates:
            if registered == 1 or step >= 5:
                continue

            try:
                created_dt = datetime.fromisoformat(created_at.replace("Z", "+00:00"))
                if created_dt.tzinfo is None:
                    created_dt = created_dt.replace(tzinfo=timezone.utc)
            except Exception:
                created_dt = now

            elapsed = now - created_dt
            current_name = name or first_name or ""

            if reminder_stage == 0 and elapsed >= REMINDER_FIRST_AFTER:
                try:
                    await context.bot.send_message(
                        chat_id=user_id,
                        text=REMINDER_1_TEXT.format(name=current_name),
                        reply_markup=continue_keyboard(),
                    )
                    db.set_reminder_stage(user_id, 1, now.isoformat())
                except Exception:
                    pass
                continue

            if reminder_stage == 1 and elapsed >= REMINDER_SECOND_AFTER:
                try:
                    await context.bot.send_message(
                        chat_id=user_id,
                        text=REMINDER_2_TEXT.format(name=current_name),
                        reply_markup=continue_keyboard(),
                    )
                    db.set_reminder_stage(user_id, 2, now.isoformat())
                except Exception:
                    pass
                continue

            if reminder_stage >= 2 and elapsed <= REMINDER_DAILY_UNTIL:
                last_dt = None
                if last_reminder_at:
                    try:
                        last_dt = datetime.fromisoformat(last_reminder_at.replace("Z", "+00:00"))
                        if last_dt.tzinfo is None:
                            last_dt = last_dt.replace(tzinfo=timezone.utc)
                    except Exception:
                        last_dt = None

                if last_dt is None or (now - last_dt) >= REMINDER_DAILY_AFTER:
                    try:
                        await context.bot.send_message(
                            chat_id=user_id,
                            text=REMINDER_3_TEXT.format(name=current_name),
                            reply_markup=continue_keyboard(),
                        )
                        db.set_reminder_stage(user_id, 2, now.isoformat())
                    except Exception:
                        pass

    except Exception as e:
        await log_exception(e)

# =========================
# ADMIN COMMANDS
# =========================
async def stats(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        if not db.is_admin(update.effective_user.id):
            await update.message.reply_text("❌ Только админ может смотреть статистику")
            return

        source = context.args[0] if context.args else None

        with db._conn() as conn:
            cur = conn.cursor()

            if source:
                cur.execute("SELECT COUNT(*) FROM users WHERE source = ?", (source,))
                total = cur.fetchone()[0]
                cur.execute("SELECT COUNT(*) FROM users WHERE source = ? AND registered = 1", (source,))
                done = cur.fetchone()[0]
                cur.execute("SELECT COUNT(*) FROM users WHERE source = ? AND registered = 0 AND step > 0", (source,))
                pending = cur.fetchone()[0]
                await update.message.reply_text(
                    f"📊 Статистика по источнику: {source}\n\nВсего: {total}\nЗавершили: {done}\nНе завершили: {pending}"
                )
            else:
                total, done, pending = db.stats()
                await update.message.reply_text(
                    f"📊 Статистика\n\nВсего: {total}\nЗавершили: {done}\nНе завершили: {pending}"
                )
    except Exception as e:
        await log_exception(e)

async def source_stats(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        if not db.is_admin(update.effective_user.id):
            await update.message.reply_text("❌ Только админ может смотреть статистику по источникам")
            return

        with db._conn() as conn:
            cur = conn.cursor()
            cur.execute("SELECT COUNT(*) FROM users")
            total = cur.fetchone()[0]
            cur.execute("""
                SELECT COALESCE(source, 'unknown') AS source, COUNT(*) AS cnt
                FROM users
                GROUP BY COALESCE(source, 'unknown')
                ORDER BY cnt DESC, source ASC
            """)
            sources = cur.fetchall()

        text = f"📊 Статистика по источникам\n\nВсего: {total}\n\n"
        for source, cnt in sources:
            text += f"• {source}: {cnt}\n"

        await update.message.reply_text(text)

    except Exception as e:
        await log_exception(e)

async def full_stats(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        if not db.is_admin(update.effective_user.id):
            await update.message.reply_text("❌ Только админ может смотреть полную статистику")
            return

        with db._conn() as conn:
            cur = conn.cursor()
            cur.execute("SELECT COUNT(*) FROM users")
            total = cur.fetchone()[0]
            cur.execute("SELECT COUNT(*) FROM users WHERE registered = 1")
            completed = cur.fetchone()[0]
            cur.execute("SELECT COUNT(*) FROM users WHERE registered = 0 AND step > 0")
            pending = cur.fetchone()[0]
            cur.execute("""
                SELECT COALESCE(source, 'unknown') AS source, COUNT(*) AS cnt
                FROM users
                GROUP BY COALESCE(source, 'unknown')
                ORDER BY cnt DESC, source ASC
            """)
            sources = cur.fetchall()
            cur.execute("""
                SELECT
                    COALESCE(source, 'unknown') AS source,
                    user_id, username, first_name, name, registered, step, created_at
                FROM users
                ORDER BY COALESCE(source, 'unknown') ASC, created_at DESC
            """)
            rows = cur.fetchall()

        text = (
            f"📊 Полная статистика\n\n"
            f"Всего: {total}\n"
            f"Завершили: {completed}\n"
            f"Не завершили: {pending}\n\n"
            f"📍 Источники:\n"
        )
        for source, cnt in sources:
            text += f"• {source}: {cnt}\n"

        text += "\n👥 Люди по источникам:\n"
        for source, user_id, username, first_name, name, registered, step, created_at in rows:
            uname = f"@{username}" if username else "-"
            display_name = name or first_name or "-"
            status = "✅" if registered == 1 else "⏳"
            text += f"\n[{source}] {status} {display_name} | {uname} | id={user_id} | step={step} | {created_at}"

        if len(text) > 3900:
            await update.message.reply_text(text[:3900])
            await update.message.reply_text(text[3900:])
        else:
            await update.message.reply_text(text)

    except Exception as e:
        await log_exception(e)

async def export_clients(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        if not db.is_admin(update.effective_user.id):
            await update.message.reply_text("❌ Только админ может выгружать базу")
            return

        file_name = "clients.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "Clients"

        headers = [
            "user_id", "username", "first_name", "name", "experience",
            "phone", "source", "registered", "step", "created_at", "updated_at",
        ]
        ws.append(headers)

        header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
        header_font = Font(color="FFFFFF", bold=True)
        header_alignment = Alignment(horizontal="center", vertical="center")

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        with sqlite3.connect(DB_FILE) as conn:
            cur = conn.cursor()
            cur.execute("""
                SELECT user_id, username, first_name, name, experience, phone,
                       source, registered, step, created_at, updated_at
                FROM users ORDER BY created_at DESC
            """)
            for row in cur.fetchall():
                ws.append(list(row))

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        for column_cells in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column_cells[0].column)
            for cell in column_cells:
                try:
                    value = "" if cell.value is None else str(cell.value)
                    if len(value) > max_length:
                        max_length = len(value)
                except Exception:
                    pass
            ws.column_dimensions[column_letter].width = min(max_length + 2, 35)

        wb.save(file_name)

        with open(file_name, "rb") as f:
            await update.message.reply_document(
                document=f,
                filename=file_name,
                caption="📁 База регистраций в Excel"
            )

        try:
            os.remove(file_name)
        except Exception:
            pass

    except Exception as e:
        await log_exception(e)
        await update.message.reply_text("❌ Ошибка при выгрузке базы")

async def add_admin(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        if not db.is_admin(update.effective_user.id):
            await update.message.reply_text("❌ Только админ может добавлять админов")
            return
        if not context.args:
            await update.message.reply_text("Использование: /add_admin <user_id>")
            return
        new_admin = int(context.args[0])
        db.add_admin(new_admin, added_by=update.effective_user.id)
        await update.message.reply_text(f"✅ Админ {new_admin} добавлен")
    except Exception as e:
        await log_exception(e)

async def remove_admin(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        if not db.is_admin(update.effective_user.id):
            await update.message.reply_text("❌ Только админ может удалять админов")
            return
        if not context.args:
            await update.message.reply_text("Использование: /remove_admin <user_id>")
            return
        admin_id = int(context.args[0])
        if admin_id == MAIN_ADMIN_ID:
            await update.message.reply_text("❌ Главного админа удалять нельзя")
            return
        if db.remove_admin(admin_id):
            await update.message.reply_text(f"✅ Админ {admin_id} удалён")
        else:
            await update.message.reply_text("❌ Не удалось удалить админа")
    except Exception as e:
        await log_exception(e)

async def admins_list(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        if not db.is_admin(update.effective_user.id):
            await update.message.reply_text("❌ Только админ может смотреть список")
            return
        rows = db.list_admins()
        if not rows:
            await update.message.reply_text("Список админов пуст")
            return
        text = "👮 Список админов:\n\n"
        for user_id, added_by, created_at in rows:
            text += f"• {user_id} | added_by={added_by} | {created_at}\n"
        await update.message.reply_text(text)
    except Exception as e:
        await log_exception(e)

# =========================
# MAIN
# =========================
def main():
    app = Application.builder().token(BOT_TOKEN).build()

    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("reminder", reminder))
    app.add_handler(CommandHandler("stats", stats))
    app.add_handler(CommandHandler("source_stats", source_stats))
    app.add_handler(CommandHandler("full_stats", full_stats))
    app.add_handler(CommandHandler("export_clients", export_clients))
    app.add_handler(CommandHandler("add_admin", add_admin))
    app.add_handler(CommandHandler("remove_admin", remove_admin))
    app.add_handler(CommandHandler("admins_list", admins_list))

    app.add_handler(CallbackQueryHandler(on_callback))
    app.add_handler(MessageHandler(filters.CONTACT, on_contact))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, on_text))

    app.job_queue.run_repeating(reminder_job, interval=REMINDER_RECHECK_EVERY_SECONDS, first=60)

    logger.info("Bot started")
    app.run_polling()

if __name__ == "__main__":
    main()
